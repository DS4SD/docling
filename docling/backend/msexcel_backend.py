import logging
from io import BytesIO
from pathlib import Path
from typing import Dict, Set, Tuple, Union

from docling_core.types.doc import (
    DoclingDocument,
    DocumentOrigin,
    GroupLabel,
    ImageRef,
    TableCell,
    TableData,
)

# from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

from docling.backend.abstract_backend import DeclarativeDocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import InputDocument

_log = logging.getLogger(__name__)

from typing import Any, List

from PIL import Image as PILImage
from pydantic import BaseModel


class ExcelCell(BaseModel):
    row: int
    col: int
    text: str
    row_span: int
    col_span: int


class ExcelTable(BaseModel):
    num_rows: int
    num_cols: int
    data: List[ExcelCell]


class MsExcelDocumentBackend(DeclarativeDocumentBackend):
    def __init__(self, in_doc: "InputDocument", path_or_stream: Union[BytesIO, Path]):
        super().__init__(in_doc, path_or_stream)

        # Initialise the parents for the hierarchy
        self.max_levels = 10

        self.parents: Dict[int, Any] = {}
        for i in range(-1, self.max_levels):
            self.parents[i] = None

        self.workbook = None
        try:
            if isinstance(self.path_or_stream, BytesIO):
                self.workbook = load_workbook(filename=self.path_or_stream)

            elif isinstance(self.path_or_stream, Path):
                self.workbook = load_workbook(filename=str(self.path_or_stream))

            self.valid = True
        except Exception as e:
            self.valid = False

            raise RuntimeError(
                f"MsPowerpointDocumentBackend could not load document with hash {self.document_hash}"
            ) from e

    def is_valid(self) -> bool:
        _log.info(f"valid: {self.valid}")
        return self.valid

    @classmethod
    def supports_pagination(cls) -> bool:
        return True

    def unload(self):
        if isinstance(self.path_or_stream, BytesIO):
            self.path_or_stream.close()

        self.path_or_stream = None

    @classmethod
    def supported_formats(cls) -> Set[InputFormat]:
        return {InputFormat.XLSX}

    def convert(self) -> DoclingDocument:
        # Parses the XLSX into a structured document model.

        origin = DocumentOrigin(
            filename=self.file.name or "file.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            binary_hash=self.document_hash,
        )

        doc = DoclingDocument(name=self.file.stem or "file.xlsx", origin=origin)

        if self.is_valid():
            doc = self._convert_workbook(doc)
        else:
            raise RuntimeError(
                f"Cannot convert doc with {self.document_hash} because the backend failed to init."
            )

        return doc

    def _convert_workbook(self, doc: DoclingDocument) -> DoclingDocument:

        if self.workbook is not None:

            # Iterate over all sheets
            for sheet_name in self.workbook.sheetnames:
                _log.info(f"Processing sheet: {sheet_name}")

                # Access the sheet by name
                sheet = self.workbook[sheet_name]

                self.parents[0] = doc.add_group(
                    parent=None,
                    label=GroupLabel.SECTION,
                    name=f"sheet: {sheet_name}",
                )

                doc = self._convert_sheet(doc, sheet)
        else:
            _log.error("Workbook is not initialized.")

        return doc

    def _convert_sheet(self, doc: DoclingDocument, sheet: Worksheet):

        doc = self._find_tables_in_sheet(doc, sheet)

        doc = self._find_images_in_sheet(doc, sheet)

        return doc

    def _find_tables_in_sheet(self, doc: DoclingDocument, sheet: Worksheet):

        tables = self._find_data_tables(sheet)

        for excel_table in tables:
            num_rows = excel_table.num_rows
            num_cols = excel_table.num_cols

            table_data = TableData(
                num_rows=num_rows,
                num_cols=num_cols,
                table_cells=[],
            )

            for excel_cell in excel_table.data:

                cell = TableCell(
                    text=excel_cell.text,
                    row_span=excel_cell.row_span,
                    col_span=excel_cell.col_span,
                    start_row_offset_idx=excel_cell.row,
                    end_row_offset_idx=excel_cell.row + excel_cell.row_span,
                    start_col_offset_idx=excel_cell.col,
                    end_col_offset_idx=excel_cell.col + excel_cell.col_span,
                    col_header=False,
                    row_header=False,
                )
                table_data.table_cells.append(cell)

            doc.add_table(data=table_data, parent=self.parents[0])

        return doc

    def _find_data_tables(self, sheet: Worksheet):
        """
        Find all compact rectangular data tables in a sheet.
        """
        # _log.info("find_data_tables")

        tables = []  # List to store found tables
        visited: set[Tuple[int, int]] = set()  # Track already visited cells

        # Iterate over all cells in the sheet
        for ri, row in enumerate(sheet.iter_rows(values_only=False)):
            for rj, cell in enumerate(row):

                # Skip empty or already visited cells
                if cell.value is None or (ri, rj) in visited:
                    continue

                # If the cell starts a new table, find its bounds
                table_bounds, visited_cells = self._find_table_bounds(
                    sheet, ri, rj, visited
                )

                visited.update(visited_cells)  # Mark these cells as visited
                tables.append(table_bounds)

        return tables

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        visited: set[Tuple[int, int]],
    ):
        """
        Determine the bounds of a compact rectangular table.
        Returns:
        - A dictionary with the bounds and data.
        - A set of visited cell coordinates.
        """
        _log.info("find_table_bounds")

        max_row = self._find_table_bottom(sheet, start_row, start_col)
        max_col = self._find_table_right(sheet, start_row, start_col)

        # Collect the data within the bounds
        data = []
        visited_cells = set()
        for ri in range(start_row, max_row + 1):
            for rj in range(start_col, max_col + 1):

                cell = sheet.cell(row=ri + 1, column=rj + 1)  # 1-based indexing

                # Check if the cell belongs to a merged range
                row_span = 1
                col_span = 1

                # _log.info(sheet.merged_cells.ranges)
                for merged_range in sheet.merged_cells.ranges:

                    if (
                        merged_range.min_row <= ri + 1
                        and ri + 1 <= merged_range.max_row
                        and merged_range.min_col <= rj + 1
                        and rj + 1 <= merged_range.max_col
                    ):

                        row_span = merged_range.max_row - merged_range.min_row + 1
                        col_span = merged_range.max_col - merged_range.min_col + 1
                        break

                if (ri, rj) not in visited_cells:
                    data.append(
                        ExcelCell(
                            row=ri - start_row,
                            col=rj - start_col,
                            text=str(cell.value),
                            row_span=row_span,
                            col_span=col_span,
                        )
                    )
                    # _log.info(f"cell: {ri}, {rj} -> {ri - start_row}, {rj - start_col}, {row_span}, {col_span}: {str(cell.value)}")

                    # Mark all cells in the span as visited
                    for span_row in range(ri, ri + row_span):
                        for span_col in range(rj, rj + col_span):
                            visited_cells.add((span_row, span_col))

        return (
            ExcelTable(
                num_rows=max_row + 1 - start_row,
                num_cols=max_col + 1 - start_col,
                data=data,
            ),
            visited_cells,
        )

    def _find_table_bottom(self, sheet: Worksheet, start_row: int, start_col: int):
        """Function to find the bottom boundary of the table"""

        max_row = start_row

        while max_row < sheet.max_row - 1:
            # Get the cell value or check if it is part of a merged cell
            cell = sheet.cell(row=max_row + 2, column=start_col + 1)

            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand max_row to include the merged range if applicable
            if merged_range:
                max_row = max(max_row, merged_range.max_row - 1)
            else:
                max_row += 1

        return max_row

    def _find_table_right(self, sheet: Worksheet, start_row: int, start_col: int):
        """Function to find the right boundary of the table"""

        max_col = start_col

        while max_col < sheet.max_column - 1:
            # Get the cell value or check if it is part of a merged cell
            cell = sheet.cell(row=start_row + 1, column=max_col + 2)

            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand max_col to include the merged range if applicable
            if merged_range:
                max_col = max(max_col, merged_range.max_col - 1)
            else:
                max_col += 1

        return max_col

    def _find_images_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet
    ) -> DoclingDocument:

        # Iterate over byte images in the sheet
        for idx, image in enumerate(sheet._images):  # type: ignore

            try:
                pil_image = PILImage.open(image.ref)

                doc.add_picture(
                    parent=self.parents[0],
                    image=ImageRef.from_pil(image=pil_image, dpi=72),
                    caption=None,
                )
            except:
                _log.error("could not extract the image from excel sheets")

        """
        for idx, chart in enumerate(sheet._charts):  # type: ignore
            try:
                chart_path = f"chart_{idx + 1}.png"
                _log.info(
                    f"Chart found, but dynamic rendering is required for: {chart_path}"
                )

                _log.info(f"Chart {idx + 1}:")
                
                # Chart type
                # _log.info(f"Type: {type(chart).__name__}")
                print(f"Type: {type(chart).__name__}")

                # Extract series data
                for series_idx, series in enumerate(chart.series):
                    #_log.info(f"Series {series_idx + 1}:")
                    print(f"Series {series_idx + 1} type: {type(series).__name__}")
                    #print(f"x-values: {series.xVal}")
                    #print(f"y-values: {series.yVal}")

                    print(f"xval type: {type(series.xVal).__name__}")
                    
                    xvals = []
                    for _ in series.xVal.numLit.pt:
                        print(f"xval type: {type(_).__name__}")
                        if hasattr(_, 'v'):
                            xvals.append(_.v)

                    print(f"x-values: {xvals}")
                            
                    yvals = []
                    for _ in series.yVal:
                        if hasattr(_, 'v'):
                            yvals.append(_.v)
                            
                    print(f"y-values: {yvals}")                    
                    
            except Exception as exc:
                print(exc)
                continue
        """

        return doc
